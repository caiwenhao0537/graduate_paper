#-*- coding:utf-8 -*-
from openpyxl import Workbook
from openpyxl import load_workbook
import csv
import pandas as pd
import numpy as np
from pandas import DataFrame,Series

# patent_data1=pd.read_excel('D:\pydata\mypydata\paper_data\plane_83.xlsx')
# patent_data_df1=DataFrame(patent_data1)
# patent_data_df1.columns=['专利号','标题','德温特标题','IPC','德温特摘要','德温特摘要-用途','权利要求项','独立权利要求项','说明书']
# patent_data_df1.drop([0],inplace=True)
# patent_data_df1.to_csv('D:\pydata\mypydata\paper_data\plane_83.csv')
# #print(type(patent_data_df))
# patent_data2=pd.read_excel('D:\pydata\mypydata\paper_data\plane_9410.xlsx')
# patent_data_df2=DataFrame(patent_data2)
# patent_data_df2.columns=['专利号','标题','德温特标题','IPC','德温特摘要','德温特摘要-用途','权利要求项','独立权利要求项','说明书']
# patent_data_df2.drop([0],inplace=True)
# patent_data_df2.to_csv('D:\pydata\mypydata\paper_data\plane_9410.csv')
#
# patent_data3=pd.read_excel('D:\pydata\mypydata\paper_data\plane_1117.xlsx')
# patent_data_df3=DataFrame(patent_data3)
# patent_data_df3.columns=['专利号','标题','德温特标题','IPC','德温特摘要','德温特摘要-用途','权利要求项','独立权利要求项','说明书']
# patent_data_df3.drop([0],inplace=True)
# patent_data_df3.to_csv('D:\pydata\mypydata\paper_data\plane_1117.csv')



#patent_data4=patent_data_df1+patent_data_df2+patent_data_df3
#patent_data4.to_csv('D:\pydata\mypydata\paper_data\plane2.csv')

#wb1 = load_workbook('D:\pydata\mypydata\paper_data\plane_83.xlsx')
#wb2 = load_workbook('D:\pydata\mypydata\paper_data\plane_9410.xlsx')
#wb3 = load_workbook('D:\pydata\mypydata\paper_data\plane_1117.xlsx')
#ws1=wb1['Sheet1']
#ws2=wb2['Sheet1']
#ws3=wb3['Sheet1']


# with open('D:\pydata\mypydata\paper_data\plane.csv','w',encoding='utf-8') as f:
#     writer=csv.writer(f)
#     writer.writerow(['专利号','标题','德温特标题','IPC','德温特摘要','德温特摘要-用途','权利要求项','独立权利要求项','说明书'])
#     data=[]
#     for row in ws1.rows:
#         content=[]
#         for col in row:
#             if col.value != '':
#                 content.append(col.value)
        #print(content)
        #data.append(content)
    #writer.writerows(data)
# for row in ws.iter_rows('A3:I3'):
#     print(row)
#f1=open('D:\pydata\mypydata\paper_data\plane_83.xlsx','r')
#f2=open('D:\pydata\mypydata\paper_data\plane_9410.xlsx','a')
#f3=open('D:\pydata\mypydata\paper_data\plane_1117.xlsx','a')


# for line in data_f2:
#     f1.write(line)
#     f1.write('\n')
# for line1 in data_f3:
#     f1.write(line1)
#     f1.write('\n')

# f1.close()
# f2.close()
# f3.close()